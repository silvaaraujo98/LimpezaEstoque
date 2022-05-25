from openpyxl import load_workbook
def limpa_linha (path_completo):
    #Função que pega o caminho de uma planilha e limpa os dados nulos

    wb=load_workbook(path_completo)
    ws=wb['Sheet1']
    i=5
    while i<= ws.max_row:
        flag = ws.cell(i,26).value
        if flag == None:
            ws.delete_rows(i,1)
            wb.save("Compras2.xlsx")
            i-=1
        i+=1
    wb.save("Compras2.xlsx")