# Importando as bibliotecas
from openpyxl import load_workbook
from limpa_linha import limpa_linha


def transforma_base(path_completo):
    
    #Chamando a função que limpa as linhas
    limpa_linha(path_completo)
    
    
    #Carregando as bases
    wb_compras = load_workbook(path_completo)
    wb_base = load_workbook("Base.xlsx")

    #Setando as planilhas
    ws_compras = wb_compras['Sheet1']
    ws_base = wb_base['Plan1']

    #Setando o restaurante
    restaurante = ws_compras.cell(1,1).value[5:]


    #Reconhecendo a linha máxima do arquivo de compras
    linha_maxima = ws_compras.max_row
    i=5
    while i<= linha_maxima:
        Lista = []
        produto = ws_compras.cell(i,9).value
        codigo = ws_compras.cell(i,3).value
        if None not in [codigo,produto]:
            j=1
            fornecedor = ws_compras.cell(i+j,7).value
            while fornecedor != None:
                data = ws_compras.cell(i+j,21).value
                qtde = ws_compras.cell(i+j,26).value
                preco_unitario = ws_compras.cell(i+j,29).value
                extensao = ws_compras.cell(i+j,33).value
                Lista = [restaurante,produto,fornecedor,data,qtde,preco_unitario,extensao]
                ws_base.append(Lista)
                j+=1
                fornecedor = ws_compras.cell(i+j,7).value
        i+=1
    wb_base.save("Base.xlsx")



